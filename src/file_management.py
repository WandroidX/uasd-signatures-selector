import openpyxl
from src import dict_management as dictma


def get_signatures_from_file(campus_where: str,  signatures_file: str, signatures_names: list[str], useful_columns_index: list[int], useful_columns_names: list[str]) -> list[dict]:
    '''Parse signatures from a excel file from the UASD. Convert rows in dicts'''

    xl_sheet = openpyxl.load_workbook(signatures_file)
    active_sheet = xl_sheet['Hoja1']
    active_sheet_max_row = active_sheet.max_row

    signatures: list[dict] = []
    SIGNATURE_NAME_INDEX: int = 3
    CAMPUS_NAME_INDEX: int = 8

    for i in range(1, active_sheet_max_row):
        current_signature_dict: dict = {}
        current_row_campus = str(active_sheet.cell(row = i, column = CAMPUS_NAME_INDEX).value)
        current_row_signature_name= str(active_sheet.cell(row = i, column = SIGNATURE_NAME_INDEX).value)
        if current_row_campus == campus_where:
            if current_row_signature_name in signatures_names:
                for x, name in zip(useful_columns_index, useful_columns_names):
                    current_cell: str = str(active_sheet.cell(row = i, column = x).value)
                    if current_cell:
                        current_signature_dict.update({name: current_cell})
                    else:
                        break

                else:
                    signatures.append(current_signature_dict)


    merged_signatures: list[dict] = []
    # here, the dicts are merged (this is in case that a signature have two differents 
    # dicts, will be combined in one)
    for signature_dict in signatures:
        

        dicts_list: list[dict] = dictma.get_dicts_from_dicts_list(signatures, {
            'NRC': signature_dict['NRC']
        }, can_be_several=True)
        merged_signatures.append(dictma.merge_dicts(dicts_list))



    return merged_signatures

def write_signatures_data(file: str, signatures_dict_list: list[dict], useful_keys: list[str]):
    with open(file=file, mode='a') as uasd_data_file:
        # delete signatures from dict list if status is closed
        open_signatures = list( filter(
            lambda element: element['Status'] != 'Sección Cerrada', signatures_dict_list
        ))
        signatures_filtered_info = []
        for dict in open_signatures :
            current_signature_dict = {}
            for (key, value) in dict.items():
                if key in useful_keys:
                    current_signature_dict.update({key: value})
            signatures_filtered_info.append(current_signature_dict)


        for i, dicti in enumerate(signatures_filtered_info):
            dicti_values = list(dicti.values())
            if i == 0:
                dicti_keys = list(dicti.keys())
                dicti_keys_str = f'{" ".join(dicti_keys)}\n'
                dicti_values[0] = dicti_keys_str + dicti_values[0]
            text = ' '.join(dicti_values)
            uasd_data_file.write(text + '\n')


